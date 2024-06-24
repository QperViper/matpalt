from tkinter import messagebox
import tkinter as tk
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from CTkMessagebox import CTkMessagebox
from openpyxl.workbook import Workbook
from customtkinter import *
from CTkListbox import *
from customtkinter import  CTkButton, CTkEntry, CTkLabel
import customtkinter as ctk
from datetime import datetime
from datetime import date
from tkcalendar import Calendar

import conexion


class Interfaz(object):


    def __init__(self) -> None:
        self.ventana=ctk.CTk()
        self.datos = conexion.Registro_de_datos()

        self.ventana.geometry("1240x620")
        self.grupo1 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo2 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo3 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo4 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo5 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo6 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo7 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo8 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo9 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        
        altura = self.ventana.winfo_reqheight()
        anchura = self.ventana.winfo_reqwidth()
        altura_pantalla = self.ventana.winfo_screenheight()
        anchura_pantalla = self.ventana.winfo_screenwidth()
        #print(f"Altura: {altura}\nAnchura: {anchura}\nAltura de pantalla: {altura_pantalla}\nAnchura de pantalla: {anchura_pantalla}")
        x = (anchura_pantalla // 5) - (anchura//4)
        y = (altura_pantalla//5) - (altura//4)
        x = (anchura_pantalla // 6) - (anchura//1)
        y = (altura_pantalla//8) - (altura//3)
        self.ventana.geometry(f"+{x}+{y}")
        # self.ventana.iconbitmap("C:\\FO_OK\\ico.ico")
        self.ventana.title("MATPALT")
        self.ventana.config(bg="green") 
        self.ventana.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.cont =0
        self.count =0
        self.items2 = []
        self.lista3 = []
        
        #HACER COPIA PARA QUE NO SE BORRE
        
        self.fecha_hoy = datetime.now()
        self.product_list=[]
        self.operaciones()
        self.opciones()
        self.ventana.mainloop()
    #listo
    def on_closing(self):
        # Aquí puedes poner el código que quieras ejecutar cuando se cierra la ventana
        print("La ventana se está cerrando.")
        if messagebox.askokcancel("Salir", "Desea salir?"):
            self.ventana.destroy()
    def opciones(self):
        
        self.btnprod = CTkButton(self.ventana,text='Agregar Gamela',width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B",command=lambda:self.agregar_nueva_fruta()).place(x=980, y=50)
        self.btnprod2 = CTkButton(self.ventana,text='Mostrar Gamela',width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B",command=lambda:self.mostrar_gamelas()).place(x=980, y=90)
        self.btntipo = CTkButton(self.ventana,text='Agregar Tipo',width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B",command=lambda:self.agregar_tipo()).place(x=1110, y=50)
        self.btntipo2 = CTkButton(self.ventana,text='Mostrar Tipo',width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',hover_color="#B8860B",bg_color='green',command=lambda:self.mostrar_tipos()).place(x=1110, y=90)
        self.btn_venta = CTkButton(self.ventana,text='Ingresar Venta',width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B",command=lambda:self.ingresar_venta()).place(x=1110, y=130)
        

    #listo
    def operaciones(self):


        self.historial_frutas = HistorialFrutas('historial_frutas.xlsx')

        
        self.lista1 = CTkListbox(self.ventana, height=400,width=480, fg_color="black", bg_color="green",font=("Arial", 14))
        self.lista1.place(x=460,y=50)

        # self.lista1.insert(0, "")

    # ========================================================
    # borra los entrys ARREGLAR ojala hacer otro para borrar el fitro y el btn 
    # ========================================================
    def borrar_widgets(self):
        self.btn_agregar2.destroy()
        self.combo.destroy()
        self.tipo2_label.destroy()
        self.kilos_label.destroy()
        self.kilos_entry.destroy()
        self.precio_label.destroy()
        self.precio_entry.destroy()
        
    def borrar_filtro(self):
        self.lista1.delete(0, tk.END)
        self.filtro_fecha.destroy()
        self.btn_filtrar.destroy()
    def borrar_widgets2(self):
        self.btn_agregar.destroy()
        self.tipo_entry.destroy()
        self.tipo_label.destroy()
        self.nom_entry.destroy()
        self.nom_label.destroy()
        
        
    # ========================================================
    # agrega tipo Arreglar 
    # ========================================================
    def agregar_tipo(self):

        try:
            self.borrar_widgets()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_filtro()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_widgets2()
            
        except:
            print("error dado pero pasado")


        self.tipo_label = CTkLabel(self.ventana,bg_color="green",text="Nombre tipo", text_color="white")
        self.tipo_label.place(x=10, y=10)

        self.tipo_entry = CTkEntry(self.ventana,bg_color="green", text_color="black")
        self.tipo_entry.place(x=10, y=50)

        self.nom_label = CTkLabel(self.ventana,bg_color="green",text="Nombre Fruta", text_color="white")
        self.nom_label.place(x=10, y=90)

        self.nom_entry = CTkEntry(self.ventana,bg_color="green", text_color="black")
        self.nom_entry.place(x=10, y=130)

        self.btn_agregar = CTkButton(self.ventana,width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B",command=self.agregar_nuevo_tipo ,text='Aceptar', text_color="white")
        self.btn_agregar.place(x=10, y=170) 
    # ========================================================
    # ejecuta el agregado tipo Arreglar (debe ingresar a una xlsx tambien)
    # ========================================================
    def agregar_nuevo_tipo(self):
        tipo = self.tipo_entry.get()
        nom = self.nom_entry.get()
        if tipo=="" or nom =="":
            self.msg = CTkMessagebox(self.ventana, title="Error", message="no existen datos para ingresar")
        else:
            try:
                tipo = self.tipo_entry.get()
                nom = self.nom_entry.get()
            except (Exception):

                self.msg2 = CTkMessagebox(self.ventana, title="Error", message="Los Kilos y el Precio deben ser numeros enteros")

        self.datos.agregar_tipo(tipo,nom)
        self.agregar_tipo()
        self.msgok = CTkMessagebox(self.ventana, title="Exito", message="datos ingresados correctamente")   
    # ========================================================
    # muestra los tipos de frutas y sus nombres
    # ========================================================
    def mostrar_tipos(self):
        consultasql = self.datos.mostrar_tipo_prod()
        result = [item[0] for item in consultasql]
        datos = [item[1] for item in consultasql]
        datos2 = [item[2] for item in consultasql]
        try:
            self.lista1.delete(0, tk.END)
            self.filtro_fecha.destroy()
            self.btn_filtrar.destroy()
        except:
            print("error pasado")
        for i in range(len(result)):
            # datoc= [datos[i],datos2[i]]
            self.lista1.insert(result[i], f"{datos[i]}, ({datos2[i]})")
            
            # self.lista1.insert(result[i],formatted_datoc[i])

    # ========================================================
    # agrega una nueva fruta que compra el dueño a un archivo xlsx
    # ========================================================
    def agregar_nueva_fruta(self):
        try:
            self.borrar_widgets()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_filtro()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_widgets2()
            
        except:
            print("error dado pero pasado")
            

        consultasql = self.datos.id_nombre_tipo()
        self.tipo2_label = CTkLabel(self.ventana,bg_color="green",text="Tipo:", text_color="white")
        self.tipo2_label.place(x=10, y=10)        
        self.kilos_label = CTkLabel(self.ventana,bg_color="green",text="Kilos:", text_color="white")
        self.kilos_label.place(x=10, y=90)
        self.kilos_entry = CTkEntry(self.ventana,bg_color="green", text_color="white")
        self.kilos_entry.place(x=10, y=130)

        self.precio_label = CTkLabel(self.ventana,bg_color="green", text="Precio", text_color="white")
        self.precio_label.place(x=10, y=170)
        self.precio_entry = CTkEntry(self.ventana,bg_color="green", text_color="white")
        self.precio_entry.place(x=10, y=210)
        
        result = [item[1] for item in consultasql]
        datos=result
        self.combo = CTkComboBox(self.ventana, bg_color="green",values=datos)
        self.combo.place(x=10, y=50)
        self.btn_agregar2 = CTkButton(self.ventana,width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B", text='Aceptar', command=self.agregar_gamela_frutas, text_color="white")
        self.btn_agregar2.place(x=10,y=250)
    # ========================================================
    # hace la ejecucion de agregar_nueva_fruta. Arreglar(debe ingresar estos datos tambien a sql)
    # ========================================================
    def agregar_gamela_frutas(self,):

        kilos = self.kilos_entry.get()
        precio = self.precio_entry.get()
        if kilos=="" or precio =="":
            self.msg = CTkMessagebox(self.ventana, title="Error", message="no existen datos para ingresar")
        else:
            try:
                
                comboint = self.combo.get()
                kilos = self.kilos_entry.get()
                precio = self.precio_entry.get()
                kilosint= int(kilos)
                precioint = int(precio)
                fecha_hoy = date.today()
                now = datetime.now()

            except (Exception):
                self.msg2 = CTkMessagebox(self.ventana, title="Error", message="Los Kilos y el Precio deben ser numeros enteros")
                print(kilosint,precioint)
            consultasql = self.datos.id_nombre_tipo()
            result = [item[1] for item in consultasql]
            print(comboint)
            
            for i in range(len(result[1])):
                for i in result:
                    self.count += 1
                    if self.combo.get() == i:
                        # result2= 
                        break                                                       
                else:
                    continue  
                break 
        self.datos.agregar_gamela_de_compra(kilosint, fecha_hoy, precioint,self.count)
        self.msgok =  CTkMessagebox(self.ventana, title="Exito", message="datos ingresados correctamente")
        self.agregar_nueva_fruta()
        self.count =0
        # self.historial_frutas.agregar_fruta(kilosint, fecha_hoy, precioint,1)
        # self.msg2 = CTkMessagebox(self.ventana, title="ok", message="ingresado")
    # ========================================================
    # ========================================================
    def mostrar_gamelas(self):
        consultasql = self.datos.mostrar_gamela()
        consultasql1 = self.datos.mostrar_tipo_prod()
        d = [item[0] for item in consultasql]
        d1 = [item[1] for item in consultasql]
        d2 = [item[2] for item in consultasql]
        d3 = [item[3] for item in consultasql]
        d4 = [item[4] for item in consultasql]
        try:
            self.filtro_fecha.destroy()
            self.btn_filtrar.destroy()
        except:
            print("error pasado")
        try:
            self.borrar_widgets()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_widgets2()
            
        except:
            print("error dado pero pasado")
            
        self.mostrar_gamelas_por_fecha()

        for i in range(0, 51):
        
            try:
                d5=int(d4[i])
                nom=self.datos.nombre_tipo(d5)
            except (Exception):
                print("error pasado")

            formatted_output = ', '.join(f"{item[0]}" for index, item in enumerate(nom))
            print(formatted_output)

            
            # datoc= [datos[i],datos2[i]]
            self.lista1.insert(d[i], f"gamela:{d[i]}, Kg:{d1[i]}, Fecha:{d2[i]}, Precio:{d3[i]}, {formatted_output}")

    def mostrar_gamelas_por_fecha(self):
        try:
            self.borrar_widgets()
            
        except:
            print("error dado pero pasado")

        # self.filtro_fecha = Calendar(self.ventana)
        # self.filtro_fecha.place(x=180,y=50)

        self.filtro_fecha = Calendar(self.ventana, locale='es_ES')
        self.filtro_fecha.place(x=180,y=50)

        self.btn_filtrar = CTkButton(self.ventana,width=120,height=30,border_width=0,corner_radius=20,fg_color='#DAA520',bg_color='green',hover_color="#B8860B", text='Filtrar por Fecha', command=self.filtro, text_color="white")
        self.btn_filtrar.place(x=180,y=250)
        
    def filtro(self):
        fecha_seleccionada = self.filtro_fecha.get_date()
        # Convertir la fecha a objeto datetime
        fecha_obj = datetime.strptime(fecha_seleccionada, '%m/%d/%y')
        # Formatear la fecha en 'YYYY-MM-DD'
        fecha_formateada = fecha_obj.strftime('%Y-%m-%d')
        # print(fecha_formateada)
        consultasql = self.datos.mostrar_gamela_por_fecha(fecha_formateada)
        consultasql1 = self.datos.id_producto(fecha_formateada)
        self.lista1.delete(0, tk.END)
        count = [item[0] for item in consultasql1]
        
        s= count[0]  # toma el valor son parentesis ni corchete

        d = [item[0] for item in consultasql]
        d1 = [item[1] for item in consultasql]
        d2 = [item[2] for item in consultasql]
        d3 = [item[3] for item in consultasql]
        d4 = [item[4] for item in consultasql]
        for i in range(0, s):
            try:
                d5=int(d4[i])
                nom=self.datos.nombre_tipo(d5)
            except (Exception):
                print("error pasado")
            formatted_output = ', '.join(f"{item[0]}" for index, item in enumerate(nom))
            print(formatted_output)
            self.lista1.insert(d[i], f"gamela:{d[i]}, Kg:{d1[i]}, Fecha:{d2[i]}, Precio:{d3[i]}, {formatted_output}")   
        




    def ingresar_venta(self):
        try:
            self.borrar_widgets()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_filtro()
            
        except:
            print("error dado pero pasado")
        try:
            self.borrar_widgets2()
            
        except:
            print("error dado pero pasado")
            

        consultasql = self.datos.id_nombre_tipo()
        self.tipo2_label = CTkLabel(self.ventana,bg_color="green",text="Tipo:", text_color="white")
        self.tipo2_label.place(x=10, y=10)        
        self.kilos_label = CTkLabel(self.ventana,bg_color="green",text="Kilos:", text_color="white")
        self.kilos_label.place(x=10, y=90)
        self.kilos_entry = CTkEntry(self.ventana,bg_color="green", text_color="white")
        self.kilos_entry.place(x=10, y=130)

        self.precio_label = CTkLabel(self.ventana,bg_color="green", text="Precio", text_color="white")
        self.precio_label.place(x=10, y=170)
        self.precio_entry = CTkEntry(self.ventana,bg_color="green", text_color="white")
        self.precio_entry.place(x=10, y=210)

        pass

    # def eliminar_pedido(self):
    #     id=CTkInputDialog(title='Eliminar producto', text='eliminar')
    #     id.geometry('300x200+800+400')
    #     dato=(int(id.get_input()))
    #     print(dato)
    #     contador_a_eliminar = [dato]  # Reemplaza con el contador del producto que deseas eliminar
    #     self.datos.eliminar_producto_por_contador(contador_a_eliminar)
    #     CTkMessagebox(title='MENSAJE', message=f'pedido {contador_a_eliminar} eliminado')
    # #listo
    def cerrar(self):
        if messagebox.askokcancel("Salir", "quieres salir?"):
            self.ventana.destroy()
        
    # def mostrar_pedidos(self):
    #     print("hola")
    #     datos = self.datos.obtener_todos_los_productos()
    #     print(datos)
    #     display_string = "\n".join([" ".join(map(str, dato)) for dato in datos])
    #     CTkMessagebox(title='base_de_datos', message=display_string)
       
#     def subir_producto_web(self):
#         dato=self.datos.traer_ultimo_id_producto()
#         for d in dato:
#             print(d)
#         if d is not None:
#     # Acceder al primer elemento de la tupla y convertirlo a un entero
#             ultimaid = d[0]
#     # Imprimir solo el número
#             print(ultimaid)
#         id_cateogoria=CTkInputDialog(title='ingrese id_categoria', text='1=Sandwiches \n 2=Pichangas \n 3=Papasfritas \n 4=Bebestibles \n 5=Colaciones \n 6=Postres \n 7=Completos \n 8=Otros')
#         id_cateogoria.geometry('500x400+600+400')
#         if ultimaid ==0 or ultimaid ==None:
#             ultimaid = 1
#         iddefault = ultimaid + 1
#         imagen=" "
# # CAMBIAR ESTOOOOOOOOOOOOOOOOOOOOOOO
#         id_cateogoria=(int(id_cateogoria.get_input()))
#         print(id_cateogoria)
#         if id_cateogoria >= 1 and id_cateogoria <=7:
#             nombre=CTkInputDialog(title='Nombre de producto', text='ingrese el nombre del producto')
#             nombre.geometry('500x400+600+400')
#             nombre=(nombre.get_input())
#             print(nombre)
#             if nombre != "":
                
#                 descripcion=CTkInputDialog(title='Descripcion del producto', text='ingrese la descripcion')
#                 descripcion.geometry('500x400+600+400')
#                 descripcion=(descripcion.get_input())
#                 print(descripcion)
#                 if descripcion != '':
                  
#                         precio=CTkInputDialog(title='Precio del producto', text='ingrese precio del producto')
#                         precio.geometry('500x400+600+400')
#                         precio=(int(precio.get_input()))
#                         print(precio)
#                         if precio > 0:
#                             self.datos.ingresar_producto_a_pagina_web(iddefault,id_cateogoria,nombre,descripcion,imagen,precio)
#                             CTkMessagebox(title='Listo!', message=f'producto= {id_cateogoria}, {descripcion}, {nombre} subido.') 
                            
#                         else:
#                             CTkMessagebox(title='Error', message=f'precio = 0, ingrese un precio mayor a 0, Operacion cancelada.')  
#                 else:
#                         CTkMessagebox(title='Error', message=f'nombre = {nombre}, ingrese un nombre, Operacion cancelada.')  
#             else:
#               CTkMessagebox(title='Error', message=f'nombre = {nombre}, ingrese un nombre, Operacion cancelada.')  
#         elif id_cateogoria >= 8:
#             CTkMessagebox(title='Error', message=f'id_categoria = {id_cateogoria}, Operacion cancelada.')
#         else:
#             CTkMessagebox(title='Error', message=f'id_categoria = {id_cateogoria}, Operacion cancelada.')
    def elim(self):
        if self.lista1.curselection() != None:
            self.lista1.delete(self.lista1.curselection())
        else:
            messagebox.showwarning("Error", "No ha seleccionado un elemento.")
    def ingreso(self,a):
        if self.lista3 == []:
            CTkMessagebox(title='Error', message='No hay productos en la lista, ingrese productos', icon="C:\\FO_OK\\ico.ico")
        print(self.valor)
        for datos in self.lista3:
            # for d in datos:
            #     print(d)
            #     # print(datos)
            #     # print (self.lista3)
            #     self.datos.ingresar_producto(d)
            contador_str,precio,palabra,fecha_como_cadena,hora_como_cadena = datos
            print(datos)
            self.datos.ingresar_producto(contador_str, precio, palabra, fecha_como_cadena, hora_como_cadena)
            dato=self.datos.traer_ultimo_id_producto()
            for d in dato:
                print(d)
            if d is not None:
    # Acceder al primer elemento de la tupla y convertirlo a un entero
                ultimaid = d[0]
    # Imprimir solo el número
            print(ultimaid)
            
            self.datos.ingresar_producto_a_boleta(contador_str,self.valor, ultimaid, fecha_como_cadena, precio)
        self.lista3 = []
        self.Eliminar_todo_de_lista()
        CTkMessagebox(title='Operacion completada', message='todo listo', icon="C:\\FO_OK\\ico.ico")
    #listo
    def insertar_elemento_en_excel(self, palabra, precio):
        
        self.lista1.insert(self.contador, palabra)
        date = self.fecha_hoy.date()
        fecha_como_cadena = self.fecha_hoy.strftime("%Y-%m-%d")
        hora = self.fecha_hoy.time()
        hora_como_cadena = self.fecha_hoy.strftime("%H:%M:%S")
        print(fecha_como_cadena)
        print(hora_como_cadena)
        self.contador += 1
        contador_str = str("'" + str(self.contador) + "'")
        
        try:
            datos = (contador_str,precio,palabra,fecha_como_cadena,hora_como_cadena)
            self.lista3.append(datos)
            #self.datos.ingresar_producto(contador_str,precio,palabra,fecha_como_cadena,hora_como_cadena)
            print("ok")
        except ValueError:
            print("eeror en ingreso de producto a mysql")
        pass
    #listo
    # Llama a la función crear_boton para crear un botón para cada elemento en 'items'
    #listo
    def crear_boton_en_pantalla(self, elemento, precio, x, y, valor):
    # Crear un botón con el elemento y el precio obtenidos
        pass
    #listo
    # def crear_boton_en_pantalla2(self, elemento, precio, x, y):
    # # Crear un botón con el elemento y el precio obtenidos
    #         btn = CTkButton(self.grupo2, text=elemento, width=180, height=30, border_width=0, corner_radius=20, command=lambda:self.insertar_elemento_en_excel(elemento, precio))
    #         btn.place(x=x, y=y)
    
    def crear_botones(self, items2, valor):
        items2 = []
        resultado_sql = self.datos.busca_id_categoria(valor)
        print(resultado_sql)
        for i, (id_producto, id_categoria, nombre, descripcion, _, precio) in enumerate(resultado_sql, start=2):
            print("Botón creado")
            print((id_producto, id_categoria, nombre, descripcion, _, precio))
            x = 10  # Posición x del botón
            y = 20 + (i - 2) * 40  # Ajusta la posición y para cada botón
            print(y)
            self.crear_boton_en_pantalla(nombre, precio, x, y, valor)
    # def crear_botones2(self, items2, valor):
    #     items2 = []
    #     resultado_sql = self.datos.busca_id_categoria(valor)
    #     print(resultado_sql)
    #     for i, (id_producto, id_categoria, nombre, descripcion, _, precio) in enumerate(resultado_sql, start=2):
    #         print("Botón creado")
    
    # # Imprimir la tupla para depurar
    #         print((id_producto, id_categoria, nombre, descripcion, _, precio))
    #         x = 10  # Posición x del botón
    #         y = 20 + (i - 2) * 40  # Ajusta la posición y para cada botón
    #         print(y)
    
    # # Llama a tu función para crear botones con los datos obtenidos
    #         self.crear_boton_en_pantalla2(nombre, precio, x, y)
    def Eliminar_todo_de_lista(self):
        self.lista1.delete(0, "end")
        
    #listo 



class HistorialFrutas:

    def __init__(self, archivo_excel):
        self.archivo_excel = archivo_excel
        self.historial = self.cargar_historial()

    def cargar_historial(self):
        try:
            workbook = load_workbook(self.archivo_excel)
            sheet = workbook.active
            historial = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                nombre, kilos, precio = row
                historial.append({'nombre': nombre, 'kilos': kilos, 'precio': precio})

            return historial
        except FileNotFoundError:
            return []

    def guardar_historial(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Nombre', 'Kilos', 'Precio'])

        for fruta in self.historial:
            sheet.append([fruta['nombre'], fruta['kilos'], fruta['precio']])

        workbook.save(self.archivo_excel)

    def agregar_fruta(self, nombre, kilos, precio):
        nueva_fruta = {'nombre': nombre, 'kilos': kilos, 'precio': precio}
        self.historial.append(nueva_fruta)
        self.guardar_historial()

    def mostrar_historial(self):
        for fruta in self.historial:
            print(f"Nombre: {fruta['nombre']}, Kilos: {fruta['kilos']}, Precio: {fruta['precio']}")

Interfaz()